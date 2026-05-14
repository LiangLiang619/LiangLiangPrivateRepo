#!/usr/bin/env python3
"""Split an .xlsx file into per-sheet CSV files (UTF-8 with BOM)."""

import csv
import os
import re
import sys


def sanitize_sheet_name(name):
    return re.sub(r'[\\/:*?"<>|]', '_', name)


def split_xlsx(xlsx_path):
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    if not os.path.isfile(xlsx_path):
        print(f"ERROR: File not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    base_name = os.path.splitext(os.path.basename(xlsx_path))[0]
    out_dir = os.path.join(os.path.dirname(xlsx_path), f"{base_name}_sheets")
    os.makedirs(out_dir, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    created_files = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        safe_name = sanitize_sheet_name(sheet_name)
        csv_path = os.path.join(out_dir, f"{safe_name}.csv")

        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                str_row = [str(cell) if cell is not None else '' for cell in row]
                if any(str_row):
                    writer.writerow(str_row)
                    row_count += 1

        if row_count > 1:
            created_files.append((csv_path, row_count - 1))
            print(f"  [OK] {safe_name}.csv  ({row_count - 1} data rows)")
        else:
            os.remove(csv_path)
            print(f"  [SKIP] {safe_name} (empty sheet)")

    wb.close()

    print(f"\nOutput directory: {out_dir}")
    print(f"Total sheets with data: {len(created_files)}")
    return out_dir, created_files


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python split_xlsx.py <input.xlsx>", file=sys.stderr)
        sys.exit(1)
    split_xlsx(sys.argv[1])
