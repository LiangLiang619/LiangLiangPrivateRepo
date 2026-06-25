"""
3C 迁移资产引用数量对比脚本

对比 LetsGo3C 迁移资产在迁移前后被其他资产引用的数量是否减少或丢失。

用法:
    python compare_ref_counts.py \
        --before "D:\backups\0617\RefByOthersAssetTable_0501.xlsx" \
        --after  "D:\backups\0617\RefByOthersAssetTable_0615.xlsx" \
        [--migration-record "...3CAssetsMigrationRecords.md"] \
        [--letsgo3c-root "E:/Dev2/LetsGoDevelop/LetsGo/Content/LetsGo3C"] \
        [--out "D:\backups\0617\3C迁移引用对比_20260617.xlsx"]
"""

import argparse
import os
import re
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass
class MigrationEntry:
    asset_name: str
    old_path: str
    new_path_record: str
    new_path_disk: Optional[str] = None
    move_type: str = ""
    owner: str = ""
    note: str = ""

    @property
    def new_path(self) -> str:
        return self.new_path_disk or self.new_path_record


@dataclass
class RefInfo:
    count: int = 0
    referrers: Set[str] = field(default_factory=set)
    referrers_complete: bool = False


def parse_migration_record(md_path: str) -> List[MigrationEntry]:
    """Parse 3CAssetsMigrationRecords.md markdown table into MigrationEntry list."""
    entries = []
    with open(md_path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    in_table = False
    header_indices = {}
    for line in lines:
        line = line.strip()
        if not line.startswith("|"):
            in_table = False
            continue

        cells = [c.strip() for c in line.split("|")]
        cells = cells[1:-1] if len(cells) > 2 else cells

        if "外部资产名" in line and "源路径" in line:
            in_table = True
            for i, h in enumerate(cells):
                header_indices[h.strip()] = i
            continue

        if not in_table:
            continue

        if all(c.replace("-", "").strip() == "" for c in cells):
            continue

        if len(cells) < 4:
            continue

        idx_name = header_indices.get("外部资产名", 1)
        idx_old = header_indices.get("源路径", 2)
        idx_new = header_indices.get("目标路径", 3)
        idx_move = header_indices.get("搬迁方式", 4)
        idx_owner = header_indices.get("负责人", 5)
        idx_note = header_indices.get("备注", -1)

        def safe_get(arr, idx, default=""):
            if idx < 0:
                return default
            return arr[idx].strip() if idx < len(arr) else default

        entry = MigrationEntry(
            asset_name=safe_get(cells, idx_name),
            old_path=safe_get(cells, idx_old),
            new_path_record=safe_get(cells, idx_new),
            move_type=safe_get(cells, idx_move),
            owner=safe_get(cells, idx_owner),
            note=safe_get(cells, idx_note),
        )
        if entry.asset_name and entry.old_path:
            entries.append(entry)

    return entries


def scan_disk_assets(letsgo3c_root: str) -> Dict[str, str]:
    """Scan LetsGo3C/Assets for .uasset/.umap, return {asset_name: /Game/LetsGo3C/...} map."""
    assets_dir = os.path.join(letsgo3c_root, "Assets")
    if not os.path.isdir(assets_dir):
        print(f"[warn] Assets directory not found: {assets_dir}")
        return {}

    content_root = os.path.dirname(letsgo3c_root)
    name_to_game_path = {}

    for root, _dirs, files in os.walk(assets_dir):
        for fname in files:
            ext = os.path.splitext(fname)[1].lower()
            if ext not in (".uasset", ".umap"):
                continue
            asset_name = os.path.splitext(fname)[0]
            full_path = os.path.join(root, fname)
            rel = os.path.relpath(full_path, content_root).replace("\\", "/")
            rel_no_ext = os.path.splitext(rel)[0]
            game_path = "/Game/" + rel_no_ext
            name_to_game_path[asset_name] = game_path

    return name_to_game_path


def resolve_new_paths(
    entries: List[MigrationEntry],
    disk_map: Dict[str, str],
) -> None:
    """For each entry, try to find the real new path on disk (handles renames)."""
    for entry in entries:
        record_target_name = entry.new_path_record.rsplit("/", 1)[-1] if entry.new_path_record else ""

        if record_target_name in disk_map:
            entry.new_path_disk = disk_map[record_target_name]
        elif entry.asset_name in disk_map:
            entry.new_path_disk = disk_map[entry.asset_name]
        else:
            entry.new_path_disk = None


def collect_lookup_paths(entries: List[MigrationEntry]) -> Set[str]:
    """Collect all old and new paths we need to look up in the xlsx files."""
    paths = set()
    for e in entries:
        if e.old_path:
            paths.add(e.old_path)
        if e.new_path_record:
            paths.add(e.new_path_record)
        if e.new_path_disk:
            paths.add(e.new_path_disk)
    return paths


def scan_xlsx(xlsx_path: str, target_paths: Set[str]) -> Dict[str, RefInfo]:
    """Stream-read xlsx, capturing only rows whose asset_path is in target_paths."""
    result: Dict[str, RefInfo] = {}
    print(f"  scanning {os.path.basename(xlsx_path)} ...", end=" ", flush=True)
    t0 = time.time()

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    row_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_count += 1
        if len(row) < 3:
            continue
        asset_path = row[1]
        if not asset_path or asset_path not in target_paths:
            continue
        ref_count = row[2] if row[2] is not None else 0
        ref_count = int(ref_count)
        referrers_str = row[3] if len(row) > 3 and row[3] else ""
        referrers = set()
        if referrers_str:
            referrers = {
                r.strip() for r in str(referrers_str).split(",")
                if r.strip() and r.strip().startswith("/Game/")
            }
        complete = (len(referrers) >= ref_count) if ref_count > 0 else True
        result[asset_path] = RefInfo(count=ref_count, referrers=referrers, referrers_complete=complete)

    wb.close()
    elapsed = time.time() - t0
    print(f"{row_count} rows in {elapsed:.1f}s, {len(result)} hits")
    return result


def determine_status(
    before_count: Optional[int],
    after_new_count: Optional[int],
    redir_count: int,
    lost_count: int,
    old_path_found: bool,
    new_path_found: bool,
) -> str:
    if not old_path_found:
        return "源路径未匹配(人工确认)"
    if not new_path_found and after_new_count is None:
        return "新路径未匹配(人工确认)"

    b = before_count if before_count is not None else 0
    a = after_new_count if after_new_count is not None else 0
    total = a + redir_count

    if total >= b and lost_count == 0:
        if total > b:
            return "正常(有新增)"
        return "正常(无减少)"
    if a < b and total >= b:
        return "减少(走重定向器兜底)"
    if total < b or lost_count > 0:
        return "疑似丢失"
    return "正常(无减少)"


HEADER = [
    "资产名",
    "旧路径(源)",
    "新路径(当前)",
    "搬迁方式",
    "负责人",
    "迁移前引用数",
    "迁移后引用数(新路径)",
    "旧路径重定向器残留",
    "迁移后合计(新+重定向)",
    "严格差值",
    "合计差值",
    "丢失引用方数",
    "丢失引用方列表",
    "新增引用方数",
    "状态",
    "备注",
]

STATUS_FILLS = {
    "正常(无减少)": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "正常(有新增)": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "减少(走重定向器兜底)": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "疑似丢失": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "源路径未匹配(人工确认)": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    "新路径未匹配(人工确认)": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}


def build_report(
    entries: List[MigrationEntry],
    before_data: Dict[str, RefInfo],
    after_data: Dict[str, RefInfo],
    out_path: str,
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资产引用对比"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    for col_idx, h in enumerate(HEADER, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    stats = {"total": 0, "ok": 0, "redir": 0, "lost": 0, "unmatched": 0}

    for row_idx, entry in enumerate(entries, 2):
        old_ref = before_data.get(entry.old_path)
        new_path = entry.new_path
        new_ref = after_data.get(new_path)
        redir_ref = after_data.get(entry.old_path)

        before_count = old_ref.count if old_ref else None
        after_new_count = new_ref.count if new_ref else None
        redir_count = redir_ref.count if redir_ref else 0

        b = before_count if before_count is not None else 0
        a = after_new_count if after_new_count is not None else 0
        total_after = a + redir_count
        delta_strict = a - b
        delta_total = total_after - b

        can_diff_refs = (
            (old_ref is not None and old_ref.referrers_complete)
            and all(
                r is None or r.referrers_complete
                for r in (new_ref, redir_ref)
            )
        )

        if can_diff_refs:
            before_referrers = old_ref.referrers if old_ref else set()
            after_all_referrers = set()
            if new_ref:
                after_all_referrers |= new_ref.referrers
            if redir_ref:
                after_all_referrers |= redir_ref.referrers
            lost_referrers = before_referrers - after_all_referrers
            gained_referrers = after_all_referrers - before_referrers
        else:
            lost_referrers = set()
            gained_referrers = set()

        status = determine_status(
            before_count, after_new_count, redir_count,
            len(lost_referrers) if can_diff_refs else 0,
            old_ref is not None,
            new_ref is not None or entry.new_path_disk is not None,
        )

        lost_list = ",".join(sorted(lost_referrers)) if lost_referrers else ""
        if not can_diff_refs and (b > 0 or a > 0):
            lost_list = "(引用方列表不完整,仅以数量判定)"

        row_data = [
            entry.asset_name,
            entry.old_path,
            new_path,
            entry.move_type,
            entry.owner,
            before_count if before_count is not None else "N/A",
            after_new_count if after_new_count is not None else "N/A",
            redir_count,
            total_after,
            delta_strict,
            delta_total,
            len(lost_referrers) if can_diff_refs else "N/A(截断)",
            lost_list,
            len(gained_referrers) if can_diff_refs else "N/A(截断)",
            status,
            entry.note,
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 13:
                cell.alignment = Alignment(wrap_text=True)

        status_fill = STATUS_FILLS.get(status)
        if status_fill:
            ws.cell(row=row_idx, column=15).fill = status_fill

        stats["total"] += 1
        if "正常" in status:
            stats["ok"] += 1
        elif "重定向器" in status:
            stats["redir"] += 1
        elif "丢失" in status:
            stats["lost"] += 1
        else:
            stats["unmatched"] += 1

    col_widths = [25, 55, 55, 18, 15, 14, 20, 18, 22, 10, 10, 12, 60, 12, 22, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(out_path)
    print(f"\n[done] 输出: {out_path}")
    print(f"  总计: {stats['total']} 个迁移资产")
    print(f"  正常: {stats['ok']}")
    print(f"  重定向器兜底: {stats['redir']}")
    print(f"  疑似丢失: {stats['lost']}")
    print(f"  未匹配(人工确认): {stats['unmatched']}")


def main():
    parser = argparse.ArgumentParser(description="3C 迁移资产引用数量对比")
    parser.add_argument("--before", required=True, help="迁移前 RefByOthersAssetTable xlsx")
    parser.add_argument("--after", required=True, help="迁移后 RefByOthersAssetTable xlsx")
    parser.add_argument(
        "--migration-record",
        default=None,
        help="3CAssetsMigrationRecords.md 路径 (默认自动在 letsgo3c-root 下查找)",
    )
    parser.add_argument(
        "--letsgo3c-root",
        default=r"E:\Dev2\LetsGoDevelop\LetsGo\Content\LetsGo3C",
        help="LetsGo3C 磁盘根目录",
    )
    parser.add_argument("--out", default=None, help="输出 xlsx 路径")
    args = parser.parse_args()

    if args.migration_record is None:
        args.migration_record = os.path.join(
            args.letsgo3c_root,
            "Migration", "AssetsMigration", "3CAssetsMigrationRecords.md",
        )

    if args.out is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = os.path.dirname(os.path.abspath(args.after))
        args.out = os.path.join(out_dir, f"3C迁移引用对比_{ts}.xlsx")

    for label, path in [
        ("迁移前 xlsx", args.before),
        ("迁移后 xlsx", args.after),
        ("迁移记录 md", args.migration_record),
    ]:
        if not os.path.isfile(path):
            print(f"[error] {label} 不存在: {path}")
            sys.exit(1)

    print("=" * 60)
    print("3C 迁移资产引用数量对比")
    print("=" * 60)

    print("\n[1/5] 解析迁移记录 ...")
    entries = parse_migration_record(args.migration_record)
    print(f"  共 {len(entries)} 条迁移资产")

    print("\n[2/5] 扫描磁盘资产 ...")
    disk_map = scan_disk_assets(args.letsgo3c_root)
    print(f"  磁盘发现 {len(disk_map)} 个 .uasset/.umap")

    resolve_new_paths(entries, disk_map)
    resolved = sum(1 for e in entries if e.new_path_disk)
    print(f"  磁盘路径匹配: {resolved}/{len(entries)}")

    print("\n[3/5] 收集待查路径 ...")
    target_paths = collect_lookup_paths(entries)
    print(f"  待查路径数: {len(target_paths)}")

    print("\n[4/5] 扫描 xlsx 文件 ...")
    before_data = scan_xlsx(args.before, target_paths)
    after_data = scan_xlsx(args.after, target_paths)

    print("\n[5/5] 生成对比报告 ...")
    build_report(entries, before_data, after_data, args.out)


if __name__ == "__main__":
    main()
